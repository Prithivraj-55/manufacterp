import frappe
from frappe import _
from frappe.utils import flt, now as frappe_now, generate_hash


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _calc_qty(pig, length, width, thickness, unit_wt, sec_qty):
    """Return per-unit primary weight using the same formula as Drawing controller."""
    if pig == "Structurals":
        if length and unit_wt and sec_qty:
            return (length / 1000.0) * unit_wt * sec_qty
    elif pig == "Plates":
        if length and width and thickness and unit_wt and sec_qty:
            return (length / 1000.0) * (width / 1000.0) * thickness * unit_wt * sec_qty
    else:
        return flt(sec_qty)
    return 0.0


def _get_file_path(file_url):
    file_doc = frappe.db.get_value(
        "File", {"file_url": file_url}, "name"
    )
    if not file_doc:
        frappe.throw(_("Attached file not found. Please re-attach."))
    return frappe.get_doc("File", file_doc).get_full_path()


def _parse_excel(file_path):
    """
    Parse BOM Excel and return (drawings_dict, raw_material_rows).

    drawings_dict  : OrderedDict  {cdn: {header fields, items: [...]}}
    raw_material_rows : list of dicts (flat, one per item row)
    """
    try:
        import openpyxl
    except ImportError:
        frappe.throw(_("openpyxl is required. Run: pip install openpyxl"))

    from collections import OrderedDict

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        frappe.throw(_("Could not open Excel file: {0}").format(str(e)))

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        frappe.throw(_("Excel file is empty."))

    header = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    col_idx = {h.lower(): i for i, h in enumerate(header)}

    def _get(row, *keys):
        for k in keys:
            i = col_idx.get(k.lower())
            if i is not None and i < len(row) and row[i] is not None:
                return row[i]
        return None

    def _sflt(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).strip().replace(",", ""))
        except Exception:
            return 0.0

    def _sstr(v):
        return str(v).strip() if v is not None else ""

    drawings = OrderedDict()

    for row in all_rows[1:]:
        cdn_raw = _get(row, "customer drawing number")
        if not cdn_raw:
            continue
        cdn = _sstr(cdn_raw)

        mat_raw = _get(row, "material code")
        if not mat_raw:
            continue
        mat_code = _sstr(mat_raw)

        if cdn not in drawings:
            fg_raw = _get(row, "fg item code", "fg item", "fg_item_code", "fg_item")
            drawings[cdn] = {
                "assembly_group": _sstr(_get(row, "assembly group")),
                "customer_drawing_number": cdn,
                "duno_mark_no": _sstr(_get(row, "duno/mark no", "duno mark no")),
                "fg_item_code": _sstr(fg_raw) if fg_raw else "",
                "total_quantity": _sflt(_get(row, "total qty")),
                "total_weight": _sflt(_get(row, "total weight (kg)", "total weight")),
                "items": [],
            }

        drawings[cdn]["items"].append({
            "customer_drawing_number": cdn,
            "item_no": _sstr(_get(row, "item no")),
            "material_code": mat_code,
            "grade": _sstr(_get(row, "grade")),
            "thickness": _sflt(_get(row, "thickness")),
            "width": _sflt(_get(row, "width")),
            "length": _sflt(_get(row, "length")),
            "sec_qty": _sflt(_get(row, "reqd raw material qty", "reqd qty", "sec_qty")),
        })

    return drawings


# ---------------------------------------------------------------------------
# Whitelisted API
# ---------------------------------------------------------------------------

@frappe.whitelist()
def parse_bom_excel(so_name):
    """
    Parse the attached BOM Excel file on a Sales Order, then bulk-insert rows
    into Table 1 (Sales Order DUNO Item) and Table 2 (Sales Order Drawing Raw
    Material). Skips drawing numbers that already have a Drawing created.
    Returns {drawing_count, item_count, warnings, skipped_count}.
    """
    so = frappe.get_doc("Sales Order", so_name)
    if not so.get("custom_bom_excel_file"):
        frappe.throw(_("Please attach a BOM Excel file before loading."))

    file_path = _get_file_path(so.custom_bom_excel_file)
    drawings = _parse_excel(file_path)

    if not drawings:
        frappe.throw(_("No valid drawing rows found in the Excel file."))

    # Drawing numbers that already have a Drawing doc created — skip on reload
    locked_dnos = set(frappe.db.sql(
        """
        SELECT DISTINCT drawing_number
        FROM `tabSales Order DUNO Item`
        WHERE parent = %s AND drawing IS NOT NULL AND drawing != ''
        """,
        so_name,
        as_list=True,
    ))
    locked_dnos = {r[0] for r in locked_dnos}

    new_drawings = {cdn: d for cdn, d in drawings.items() if cdn not in locked_dnos}
    skipped_count = len(drawings) - len(new_drawings)

    if not new_drawings:
        return {
            "drawing_count": 0,
            "item_count": 0,
            "warnings": [_("{0} drawing(s) already have drawings created and were skipped.").format(skipped_count)],
            "skipped_count": skipped_count,
        }

    # --- Validate item codes (warn, don't throw) ---
    all_fg = {d["fg_item_code"] for d in new_drawings.values() if d["fg_item_code"]}
    all_mat = {i["material_code"] for d in new_drawings.values() for i in d["items"] if i["material_code"]}
    existing_items = set(frappe.db.get_all(
        "Item", filters={"name": ["in", list(all_fg | all_mat)]}, pluck="name"
    )) if (all_fg | all_mat) else set()

    warnings = []
    missing_fg = sorted(all_fg - existing_items)
    missing_mat = sorted(all_mat - existing_items)
    if missing_fg:
        warnings.append(_("FG Item codes not in Item master: {0}").format(", ".join(missing_fg)))
    if missing_mat:
        warnings.append(_("Material codes not in Item master: {0}").format(", ".join(missing_mat)))
    if skipped_count:
        warnings.append(_("{0} drawing(s) already created — skipped on reload.").format(skipped_count))

    # --- Fetch item master data for qty calculation ---
    item_data_map = {}
    if all_mat:
        for item in frappe.db.get_all(
            "Item",
            filters={"name": ["in", list(all_mat)]},
            fields=[
                "name", "item_name", "item_group", "custom_unit_weight",
                "custom_parent_item_group", "custom_secondary_uom", "stock_uom",
            ],
        ):
            item_data_map[item.name] = item

    # --- Warn for Plates missing thickness / Structurals missing unit weight in Excel ---
    dim_warn = []
    for cdn, d in new_drawings.items():
        for item in d["items"]:
            idata = item_data_map.get(item["material_code"]) or frappe._dict()
            pig = (idata.get("custom_parent_item_group") or "").strip()
            if pig == "Plates" and not flt(item.get("thickness")):
                dim_warn.append(_("{0} / {1}: Plates item missing Thickness in Excel").format(cdn, item["material_code"]))
            elif pig == "Structurals" and not flt(idata.get("custom_unit_weight")):
                dim_warn.append(_("{0} / {1}: Structurals item missing Unit Weight in Item master").format(cdn, item["material_code"]))
    if dim_warn:
        warnings.extend(dim_warn)

    # --- Fetch FG item names ---
    fg_name_map = {}
    if all_fg:
        for item in frappe.db.get_all(
            "Item",
            filters={"name": ["in", list(all_fg)]},
            fields=["name", "item_name"],
        ):
            fg_name_map[item.name] = item.item_name

    # --- Clear existing unlocked rows and reset verification flag ---
    frappe.db.sql(
        "DELETE FROM `tabSales Order Drawing Raw Material` WHERE parent = %s AND is_locked = 0",
        so_name,
    )
    frappe.db.sql(
        "DELETE FROM `tabSales Order DUNO Item` WHERE parent = %s AND (drawing IS NULL OR drawing = '')",
        so_name,
    )
    frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 0)

    # --- Re-index after delete: get next available idx for each table ---
    t1_max = frappe.db.sql(
        "SELECT COALESCE(MAX(idx), 0) FROM `tabSales Order DUNO Item` WHERE parent = %s", so_name
    )[0][0]
    t2_max = frappe.db.sql(
        "SELECT COALESCE(MAX(idx), 0) FROM `tabSales Order Drawing Raw Material` WHERE parent = %s", so_name
    )[0][0]
    t1_idx = int(t1_max) + 1
    t2_idx = int(t2_max) + 1

    now = frappe_now()
    user = frappe.session.user

    # --- Build Table 1 (Drawing List) insert values ---
    t1_fields = [
        "name", "parent", "parenttype", "parentfield", "idx",
        "creation", "modified", "modified_by", "owner", "docstatus",
        "assembly_group", "item", "item_name", "duno_mark_no", "drawing_number",
        "total_quantity", "total_weight",
        "create_drawing", "submit_drawing", "mark_final_revision", "create_bom",
    ]
    t1_values = []
    for cdn, d in new_drawings.items():
        fg = d["fg_item_code"]
        t1_values.append((
            generate_hash(length=10),
            so_name, "Sales Order", "custom_duno_items", t1_idx,
            now, now, user, user, 0,
            d["assembly_group"], fg, fg_name_map.get(fg, ""),
            d["duno_mark_no"], cdn,
            d["total_quantity"], d["total_weight"],
            1, 1, 1, 1,
        ))
        t1_idx += 1

    # --- Build Table 2 (Raw Materials) insert values ---
    t2_fields = [
        "name", "parent", "parenttype", "parentfield", "idx",
        "creation", "modified", "modified_by", "owner", "docstatus",
        "customer_drawing_number", "item_no", "material_code", "material_name",
        "item_group", "parent_item_group",
        "grade", "thickness", "width", "length",
        "sec_qty", "sec_uom", "total_sec_qty", "unit_weight", "qty", "uom", "total_weight", "is_locked",
    ]
    t2_values = []
    for cdn, d in new_drawings.items():
        tq = flt(d["total_quantity"]) or 1.0
        for item in d["items"]:
            idata = item_data_map.get(item["material_code"]) or frappe._dict()
            pig = (idata.get("custom_parent_item_group") or "").strip()
            unit_wt = flt(idata.get("custom_unit_weight") or 0)
            sec_qty = flt(item["sec_qty"])
            qty = _calc_qty(pig, item["length"], item["width"], item["thickness"], unit_wt, sec_qty)
            total_sec_qty = flt(sec_qty * tq, 3)
            total_weight = flt(qty * tq, 3)
            t2_values.append((
                generate_hash(length=10),
                so_name, "Sales Order", "custom_so_raw_materials", t2_idx,
                now, now, user, user, 0,
                cdn,
                item["item_no"],
                item["material_code"],
                idata.get("item_name") or item["material_code"],
                idata.get("item_group") or "",
                pig,
                item["grade"],
                flt(item["thickness"], 3), flt(item["width"], 3), flt(item["length"], 3),
                flt(sec_qty, 3),
                idata.get("custom_secondary_uom") or "",
                flt(total_sec_qty, 3),
                flt(unit_wt, 6),
                flt(qty, 3),
                idata.get("stock_uom") or "",
                flt(total_weight, 3),
                0,
            ))
            t2_idx += 1

    # --- Bulk insert ---
    _bulk_insert("tabSales Order DUNO Item", t1_fields, t1_values)
    _bulk_insert("tabSales Order Drawing Raw Material", t2_fields, t2_values)
    frappe.db.commit()

    return {
        "drawing_count": len(t1_values),
        "item_count": len(t2_values),
        "warnings": warnings,
        "skipped_count": skipped_count,
    }


def _bulk_insert(table, fields, values, chunk_size=200):
    """Insert rows in chunks via raw SQL for performance with 1500+ rows."""
    if not values:
        return
    fields_sql = ", ".join([f"`{f}`" for f in fields])
    placeholders = "(" + ", ".join(["%s"] * len(fields)) + ")"

    for i in range(0, len(values), chunk_size):
        chunk = values[i : i + chunk_size]
        values_sql = ", ".join([placeholders] * len(chunk))
        flat = [v for row in chunk for v in row]
        frappe.db.sql(
            f"INSERT INTO `{table}` ({fields_sql}) VALUES {values_sql}",
            flat,
        )


# ---------------------------------------------------------------------------

@frappe.whitelist()
def create_drawings_from_import(so_name):
    """
    Create Draft Drawing documents from Table 1 rows where create_drawing=1
    and drawing is blank. Pre-populates Drawing.items from matching Table 2 rows.
    Locks those Table 2 rows after creation.
    Returns list of created Drawing names.
    """
    so = frappe.get_doc("Sales Order", so_name)

    pending_rows = [
        r for r in (so.custom_duno_items or [])
        if r.create_drawing and not r.drawing
    ]
    if not pending_rows:
        frappe.throw(_("No pending rows to create drawings for."))

    # Index raw material rows by drawing number
    rm_by_cdn = {}
    for r in (so.custom_so_raw_materials or []):
        rm_by_cdn.setdefault(r.customer_drawing_number, []).append(r)

    # Validate: all material codes must exist
    all_mat = {
        r.material_code
        for rows in rm_by_cdn.values()
        for r in rows
        if r.material_code
    }
    if all_mat:
        existing = set(frappe.db.get_all("Item", filters={"name": ["in", list(all_mat)]}, pluck="name"))
        missing = sorted(all_mat - existing)
        if missing:
            frappe.throw(
                _("Cannot create drawings — material codes not in Item master:<br>{0}").format(
                    "<br>".join(missing)
                )
            )

    results = []

    for dr in pending_rows:
        cdn = dr.drawing_number
        result = {"drawing_number": cdn, "drawing": None, "status": "error", "error": ""}

        try:
            rm_rows = rm_by_cdn.get(cdn, [])

            item_data = {}
            if dr.item:
                item_data = frappe.db.get_value(
                    "Item", dr.item, ["item_name", "description"], as_dict=True
                ) or {}

            drawing = frappe.get_doc({
                "doctype": "Drawing",
                "sales_order": so_name,
                "customer": so.customer,
                "customer_name": so.customer_name,
                "customer_no": so.customer,
                "project": so.get("project"),
                "cust_po_no": so.get("po_no"),
                "fg_item_code": dr.item or "",
                "fg_item_name": item_data.get("item_name") or "",
                "fg_description": item_data.get("description") or "",
                "no_of_qty_to_manufacture": flt(dr.total_quantity),
                "duno_mark_no": dr.duno_mark_no or "",
                "customer_drawing_number": cdn or "",
                "customer_provided_wt": flt(dr.total_weight),
                "status": "Working",
            })

            no_of_qty = flt(dr.total_quantity) or 1

            for rm in rm_rows:
                pig = rm.parent_item_group or ""
                unit_wt = flt(rm.unit_weight)
                sec_qty = flt(rm.sec_qty)
                qty = _calc_qty(pig, flt(rm.length), flt(rm.width), flt(rm.thickness), unit_wt, sec_qty)
                total_qty = flt(qty) * no_of_qty
                total_sec_qty = sec_qty * no_of_qty

                drawing.append("items", {
                    "item_number": rm.item_no or "",
                    "material_code": rm.material_code,
                    "material_name": rm.material_name or "",
                    "item_group": rm.item_group or "",
                    "parent_item_group": pig,
                    "thickness": flt(rm.thickness, 3),
                    "length": flt(rm.length, 3),
                    "width": flt(rm.width, 3),
                    "sec_qty": flt(sec_qty, 3),
                    "sec_uom": rm.sec_uom or "",
                    "unit_weight": flt(unit_wt, 6),
                    "qty": flt(qty, 3),
                    "uom": rm.uom or "",
                    "total_sec_qty": flt(total_sec_qty, 3),
                    "total_qty": flt(total_qty, 3),
                })

            drawing.insert(ignore_permissions=True)
            result["drawing"] = drawing.name
            result["status"] = "success"

            # Link drawing back to Table 1 row
            frappe.db.set_value("Sales Order DUNO Item", dr.name, "drawing", drawing.name)

            # Lock matching Table 2 rows
            frappe.db.sql(
                """
                UPDATE `tabSales Order Drawing Raw Material`
                SET is_locked = 1
                WHERE parent = %s AND customer_drawing_number = %s
                """,
                (so_name, cdn),
            )
            frappe.db.commit()

        except Exception as e:
            frappe.db.rollback()
            frappe.local.message_log = []
            result["error"] = str(e)

        results.append(result)

    return results


# ---------------------------------------------------------------------------

@frappe.whitelist()
def process_drawings(so_name, step):
    """
    Run a single pipeline step for all qualifying Table 1 rows.

    step values:
      "submit"         – submit Draft drawings where submit_drawing = 1
      "final_revision" – mark submitted drawings as Final Revision where mark_final_revision = 1
      "create_bom"     – create BOM for Final Revision drawings where create_bom = 1
    """
    from manufyxinvenzaerp.drawing_management.drawing_utils import (
        mark_as_final_revision,
        create_bom_from_drawing,
    )

    so = frappe.get_doc("Sales Order", so_name)
    results = []

    for dr in (so.custom_duno_items or []):
        if not dr.drawing:
            continue

        result = {
            "drawing": dr.drawing,
            "drawing_number": dr.drawing_number or "",
            "status": "skipped",
            "detail": "",
        }

        try:
            drawing_doc = frappe.get_doc("Drawing", dr.drawing)

            if step == "submit":
                if not dr.submit_drawing:
                    result["status"] = "unchecked"
                elif drawing_doc.docstatus != 0:
                    result["status"] = "already_done"
                else:
                    drawing_doc.submit()
                    result["status"] = "success"
                    result["detail"] = "submitted"

            elif step == "final_revision":
                if not dr.mark_final_revision:
                    result["status"] = "unchecked"
                elif drawing_doc.docstatus != 1:
                    result["status"] = "skipped"
                    result["detail"] = "not submitted"
                elif drawing_doc.status == "Final Revision":
                    result["status"] = "already_done"
                else:
                    mark_as_final_revision(dr.drawing)
                    result["status"] = "success"
                    result["detail"] = "marked final revision"

            elif step == "create_bom":
                if not dr.create_bom:
                    result["status"] = "unchecked"
                elif drawing_doc.docstatus != 1 or drawing_doc.status != "Final Revision":
                    result["status"] = "skipped"
                    result["detail"] = "not in Final Revision"
                else:
                    bom_name = create_bom_from_drawing(dr.drawing)
                    result["status"] = "success"
                    result["detail"] = "bom:{0}".format(bom_name)

            elif step == "submit_bom":
                bom_name = frappe.db.get_value(
                    "BOM", {"custom_drawing": dr.drawing, "docstatus": 0}, "name"
                )
                if not bom_name:
                    result["status"] = "skipped"
                    result["detail"] = "no draft BOM"
                else:
                    bom_doc = frappe.get_doc("BOM", bom_name)
                    bom_doc.submit()
                    result["status"] = "success"
                    result["detail"] = "bom submitted: {0}".format(bom_name)

            else:
                frappe.throw(_("Unknown step: {0}").format(step))

        except Exception as e:
            frappe.db.rollback()
            frappe.local.message_log = []  # prevent throw() messages surfacing as a popup
            result["status"] = "error"
            result["error"] = str(e)

        results.append(result)

    return results


# ---------------------------------------------------------------------------

@frappe.whitelist()
def verify_raw_materials(so_name):
    """
    Validate all unlocked Raw Material rows on the Sales Order.
    Sets custom_raw_materials_verified = 1 if no issues found.
    Returns {issues: [...], verified: bool}.
    """
    so = frappe.get_doc("Sales Order", so_name)
    unlocked = [r for r in (so.custom_so_raw_materials or []) if not r.is_locked]

    if not unlocked:
        frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 1, update_modified=False)
        frappe.db.commit()
        modified = frappe.db.get_value("Sales Order", so_name, "modified")
        return {"issues": [], "verified": True, "modified": str(modified)}

    all_mat = {r.material_code for r in unlocked if r.material_code}
    existing = set(frappe.db.get_all(
        "Item", filters={"name": ["in", list(all_mat)]}, pluck="name"
    )) if all_mat else set()

    issues = []
    for row in unlocked:
        cdn = row.customer_drawing_number or "?"
        mat = row.material_code or ""
        pig = row.parent_item_group or ""

        if not mat:
            issues.append(_("Drawing {0}, Item {1}: Material Code is missing").format(cdn, row.item_no or "?"))
            continue

        if mat not in existing:
            issues.append(_("Drawing {0} / {1}: Not found in Item master").format(cdn, mat))
            continue

        if pig == "Plates":
            missing = []
            if not flt(row.thickness): missing.append("Thickness")
            if not flt(row.width):     missing.append("Width")
            if not flt(row.length):    missing.append("Length")
            if not flt(row.unit_weight): missing.append("Unit Weight (check Item master)")
            if missing:
                issues.append(_("Drawing {0} / {1} (Plates): Missing — {2}").format(cdn, mat, ", ".join(missing)))

        elif pig == "Structurals":
            missing = []
            if not flt(row.length):      missing.append("Length")
            if not flt(row.unit_weight): missing.append("Unit Weight (check Item master)")
            if missing:
                issues.append(_("Drawing {0} / {1} (Structurals): Missing — {2}").format(cdn, mat, ", ".join(missing)))

    verified = len(issues) == 0
    frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 1 if verified else 0, update_modified=False)
    frappe.db.commit()
    modified = frappe.db.get_value("Sales Order", so_name, "modified")
    return {"issues": issues, "verified": verified, "modified": str(modified)}


@frappe.whitelist()
def download_bom_template():
    """Return a pre-filled BOM Import Excel template as a file download."""
    try:
        import openpyxl
    except ImportError:
        frappe.throw(_("openpyxl is required. Run: pip install openpyxl"))

    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM Import"

    headers = [
        "Assembly Group", "Customer Drawing Number", "DUNO/Mark No",
        "FG Item", "Total Qty", "Total Weight (KG)",
        "Item No", "Material Code", "Grade", "Thickness", "Width", "Length",
        "Reqd Raw Material Qty",
    ]
    ws.append(headers)

    # Sample row 1 — drawing CDN-001, item 1
    ws.append([
        "Structural Assembly", "CDN-001", "DM-001", "FG-ITEM-001", 5, 250.0,
        "1", "MAT-STRUCT-001", "A36", 0, 0, 3000, 2,
    ])
    # Sample row 2 — same drawing CDN-001, item 2 (same header columns repeated)
    ws.append([
        "Structural Assembly", "CDN-001", "DM-001", "FG-ITEM-001", 5, 250.0,
        "2", "MAT-PLATE-001", "IS2062", 10, 200, 1500, 1,
    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "BOM_Import_Template.xlsx"
    frappe.response["filecontent"] = output.read()
    frappe.response["type"] = "download"


@frappe.whitelist()
def clear_drawing_import(so_name):
    """
    Remove the BOM Excel attachment and delete all unlocked import rows.
    Rows that already have a Drawing created (locked) are preserved.
    Returns counts of deleted rows.
    """
    # Count what will be removed
    t1_del = frappe.db.sql(
        "SELECT COUNT(*) FROM `tabSales Order DUNO Item` WHERE parent = %s AND (drawing IS NULL OR drawing = '')",
        so_name, as_list=True
    )[0][0]
    t2_del = frappe.db.sql(
        "SELECT COUNT(*) FROM `tabSales Order Drawing Raw Material` WHERE parent = %s AND is_locked = 0",
        so_name, as_list=True
    )[0][0]

    # Delete unlocked rows
    frappe.db.sql(
        "DELETE FROM `tabSales Order DUNO Item` WHERE parent = %s AND (drawing IS NULL OR drawing = '')",
        so_name,
    )
    frappe.db.sql(
        "DELETE FROM `tabSales Order Drawing Raw Material` WHERE parent = %s AND is_locked = 0",
        so_name,
    )

    # Clear the file attachment field and verification flag on the SO
    frappe.db.set_value("Sales Order", so_name, "custom_bom_excel_file", "")
    frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 0)
    frappe.db.commit()

    return {"deleted_drawings": int(t1_del), "deleted_items": int(t2_del)}

"""T5 — Nature of Work and Rate Schedule ride in on the BOM sheet and are validated
against their masters when Raw Materials are verified.

Validated by record NAME only, with no format rule: Rate Schedule is named by its own
RS No, so the name IS the title being typed ("RS- O/S-001 A"). A pattern guessed from
one example would start rejecting valid codes as soon as the client's numbering
changed.

Covers the whole path: the download template carries the columns, the parser reads
them, a value that is not in its master blocks verification and names the drawing it
came from, and correcting it lets verification through.

Run: bench --site manufact execute manufyxinvenzaerp.tests.verify_bom_nature_of_work_rate_schedule.run
"""

import io
import frappe

checks = []


def check(label, got, want):
    ok = got == want
    checks.append(ok)
    print("  %-4s %-62s got=%r want=%r" % ("OK" if ok else "FAIL", label, got, want))


def _template_headers():
    from manufyxinvenzaerp.drawing_management.so_drawing_import import download_bom_template
    import openpyxl
    frappe.response.clear()
    download_bom_template()
    wb = openpyxl.load_workbook(io.BytesIO(frappe.response["filecontent"]), data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    frappe.response.clear()
    return rows


def _parse_sheet(rows):
    """Write rows to a temp xlsx and run the real parser over it."""
    from manufyxinvenzaerp.drawing_management.so_drawing_import import _parse_excel
    import openpyxl, tempfile, os
    wb = openpyxl.Workbook()
    for r in rows:
        wb.active.append(list(r))
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    try:
        return _parse_excel(path)
    finally:
        os.unlink(path)


def run():
    from manufyxinvenzaerp.drawing_management.so_drawing_import import _check_drawing_masters

    real_now = frappe.db.get_value("Nature of Work", {}, "name")
    real_rs = frappe.db.get_value("Rate Schedule", {}, "name")
    print("masters on this site: Nature of Work=%r  Rate Schedule=%r" % (real_now, real_rs))

    print()
    print("=== download template carries both columns ===")
    rows = _template_headers()
    header = list(rows[0])
    check("Nature of Work column present", "Nature of Work" in header, True)
    check("Rate Schedule column present", "Rate Schedule" in header, True)
    sample = rows[1]
    check("sample row is pre-filled from the real masters",
          (sample[header.index("Nature of Work")], sample[header.index("Rate Schedule")]),
          (real_now, real_rs))

    print()
    print("=== parser reads them off the sheet ===")
    body = [
        ["Structural Assembly", "CDN-T5", "DM-T5", "FG-T5", 1, 100.0,
         real_now, real_rs, "1", "MAT-T5", "A36", 0, 0, 3000, 1],
    ]
    parsed = _parse_sheet([header] + body)
    d = parsed.get("CDN-T5") or {}
    check("nature_of_work parsed", d.get("nature_of_work"), real_now)
    check("rate_schedule parsed", d.get("rate_schedule"), real_rs)

    print()
    print("=== a value not in the master blocks verification ===")
    so = frappe._dict(custom_duno_items=[
        frappe._dict(drawing_number="CDN-GOOD", duno_mark_no="DM-1",
                     nature_of_work=real_now, rate_schedule=real_rs),
        frappe._dict(drawing_number="CDN-BAD", duno_mark_no="DM-2",
                     nature_of_work=real_now, rate_schedule="RS- O/S-999 Z"),
    ])
    so.get = lambda f, *a, **k: so[f] if f in so else None
    issues = _check_drawing_masters(so)
    check("exactly one issue raised", len(issues), 1)
    check("it names the offending drawing", "CDN-BAD" in (issues[0] if issues else ""), True)
    check("it names the offending value", "RS- O/S-999 Z" in (issues[0] if issues else ""), True)
    check("the good drawing is not flagged", any("CDN-GOOD" in i for i in issues), False)
    if issues:
        print("       ->", frappe.utils.strip_html(issues[0])[:140])

    print()
    print("=== correcting it lets verification through ===")
    so.custom_duno_items[1].rate_schedule = real_rs
    check("no issues after correction", _check_drawing_masters(so), [])

    print()
    print("=== an unknown Nature of Work is caught the same way ===")
    so.custom_duno_items[1].nature_of_work = "Not A Real Nature"
    issues = _check_drawing_masters(so)
    check("caught", len(issues), 1)
    check("named as Nature of Work", "Nature of Work" in (issues[0] if issues else ""), True)

    print()
    print("=== blank is allowed (neither is mandatory, and old imports predate them) ===")
    so.custom_duno_items[1].nature_of_work = ""
    so.custom_duno_items[1].rate_schedule = ""
    check("blank passes", _check_drawing_masters(so), [])

    print()
    print("=== SUMMARY ===")
    if all(checks):
        print("ALL %d CHECKS PASSED" % len(checks))
    else:
        print("%d of %d CHECKS FAILED" % (checks.count(False), len(checks)))
